from __future__ import annotations

from io import BytesIO
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from SplitWorkload.backend.app.models.api import AnalysisResponse, SheetResult

_ROLES: Iterable[str] = ("product", "frontend", "backend", "test", "ops")


class ExcelExporter:
    """Build an Excel workbook from the analysis response."""

    def build_workbook(self, response: AnalysisResponse) -> bytes:
        workbook = Workbook()
        # remove default sheet to avoid empty tab
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        for index, sheet in enumerate(response.sheets, start=1):
            title = sheet.sheet_name.strip() or f"Sheet{index}"
            worksheet = workbook.create_sheet(title=title[:31])
            self._write_sheet(worksheet, sheet)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _write_sheet(self, worksheet, sheet: SheetResult) -> None:
        worksheet.append(["Sheet 名称", sheet.sheet_name])
        worksheet.append([
            "总计人月",
            _format_decimal(sheet.summary.total_allocated),
            "限制",
            _format_decimal(sheet.summary.limit) if sheet.summary.limit is not None else "",
        ])
        worksheet.append([])

        worksheet.append(["角色", "人月"])
        for role, value in sheet.summary.by_role.items():
            worksheet.append([role, _format_decimal(value)])
        worksheet.append([])

        worksheet.append([
            "项目/模块",
            "业务需求说明",
            "产品",
            "前端",
            "后端",
            "测试",
            "运维",
            "分析说明",
        ])

        for project in sheet.projects:
            allocation = project.allocation
            worksheet.append(
                [
                    project.project or "",
                    project.requirement,
                    _format_decimal(allocation.product),
                    _format_decimal(allocation.frontend),
                    _format_decimal(allocation.backend),
                    _format_decimal(allocation.test),
                    _format_decimal(allocation.ops),
                    allocation.analysis or "",
                ]
            )

        self._auto_size_columns(worksheet)

    def _auto_size_columns(self, worksheet) -> None:
        for column_cells in worksheet.columns:
            try:
                column = get_column_letter(column_cells[0].column)
            except AttributeError:
                continue
            max_length = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column].width = min(max_length + 2, 60)


def _format_decimal(value: Optional[float]) -> str:
    if value is None:
        return ""
    return f"{float(value):.1f}"
